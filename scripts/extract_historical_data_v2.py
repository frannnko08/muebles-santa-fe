#!/usr/bin/env python3
"""
Script MEJORADO para extraer datos históricos de archivos Excel.
VERSIÓN 2.0 - Busca recursivamente en TODAS las subcarpetas

Uso:
    python3 extract_historical_data.py /ruta/a/archivos_historicos/ datos_extraidos.csv

Estructura soportada:
    archivos_historicos/
    ├── COTIZACIONES/
    │   ├── 2026-Enero/
    │   │   ├── archivo.xlsx
    │   ├── 2026-Febrero/
    │   └── ...
    └── NOTAS_VENTA/
        ├── 2026-Enero/
        ├── 2026-Febrero/
        └── ...
        
O simplemente:
    archivos_historicos/
    ├── 2026-Enero/
    ├── 2026-Febrero/
    └── ...
"""

import openpyxl
import csv
import sys
from pathlib import Path
from datetime import datetime, timedelta
import json
import re

def excel_date_to_iso(excel_date):
    """Convierte número de fecha Excel a formato ISO (YYYY-MM-DD)"""
    if isinstance(excel_date, str):
        try:
            excel_date = int(float(excel_date))
        except (ValueError, TypeError):
            return None
    
    if isinstance(excel_date, (int, float)):
        try:
            base_date = datetime(1899, 12, 30)
            result_date = base_date + timedelta(days=excel_date)
            return result_date.strftime("%Y-%m-%d")
        except (OverflowError, ValueError):
            return None
    
    if hasattr(excel_date, 'date'):  # datetime object
        return excel_date.strftime("%Y-%m-%d")
    
    return None

def extract_from_resumen_sheet(wb, filename):
    """Extrae datos de la pestaña RESUMEN (si existe)
    PERO busca NETO en la hoja principal para multiplicar por 1.19"""
    try:
        ws = wb["RESUMEN"]
    except KeyError:
        return None
    
    data = []
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    
    if len(rows) < 2:
        return None
    
    # Obtener headers (primera fila)
    headers = [str(h or "").strip().upper() for h in rows[0]]
    
    # Detectar tipo de documento
    is_nota_venta = "NOTA DE VENTA" in headers
    
    # Procesar datos (desde fila 2)
    for row in rows[1:]:
        if not any(row):
            continue
        
        try:
            if is_nota_venta:
                # Estructura: COTIZACION | NOTA DE VENTA | NOMBRE | FECHA | TOTAL
                if len(row) >= 5:
                    cotizacion = str(row[0]).strip() if row[0] else ""
                    nota_venta = str(row[1]).strip() if row[1] else ""
                    cliente = str(row[2]).strip() if row[2] else ""
                    fecha_val = row[3]
                    total_val = row[4]  # Valor del RESUMEN (puede ser NETO o TOTAL)
                    tipo = "nota_venta"
                else:
                    continue
            else:
                # Estructura: N° COTIZACION | CLIENTE | FECHA | TOTAL
                if len(row) >= 4:
                    cotizacion = str(row[0]).strip() if row[0] else ""
                    cliente = str(row[1]).strip() if row[1] else ""
                    fecha_val = row[2]
                    total_val = row[3]
                    nota_venta = ""
                    tipo = "cotizacion"
                else:
                    continue
            
            # Convertir fecha
            if hasattr(fecha_val, 'date'):  # datetime object
                fecha = fecha_val.strftime("%Y-%m-%d")
            else:
                fecha = excel_date_to_iso(fecha_val)
            
            # IMPORTANTE: El valor en RESUMEN es NETO
            # Para BARRANES: no se multiplica por 1.19 (sin IVA)
            # Para NOTAS/COTIZACIONES: se multiplica por 1.19
            total = float(total_val) if total_val and isinstance(total_val, (int, float)) else 0
            if total > 0:
                total = total * 1  # Barranes NO tienen IVA
            
            if cotizacion and cliente and fecha and total > 0:
                data.append({
                    "numero_cotizacion": cotizacion.replace("N° ", "").strip(),
                    "numero_nota_venta": nota_venta,
                    "cliente": cliente,
                    "fecha": fecha,
                    "total": total,
                    "tipo": tipo,
                    "canal": "web"
                })
        
        except (IndexError, ValueError, TypeError, AttributeError):
            continue
    
    return data if data else None

def find_data_sheet(wb):
    """Encuentra la hoja que contiene los datos principales"""
    for sheet_name in wb.sheetnames:
        if sheet_name in ["RESUMEN", "REGLAS", "COSTOS", "CUBIERTA", "NOTA DE VENTA", "PRODUCCION", "SEGUIMIENTO"]:
            continue
        ws = wb[sheet_name]
        text = ""
        for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
            text += " ".join(str(v or "") for v in row).upper()
        if "PRESUPUESTO" in text or "CLIENTE" in text or "NOMBRE" in text:
            return ws, sheet_name
    
    # Si no encontró, busca el primer sheet que no sea especial
    for sheet_name in wb.sheetnames:
        if sheet_name not in ["RESUMEN", "REGLAS", "COSTOS", "CUBIERTA", "NOTA DE VENTA", "PRODUCCION", "SEGUIMIENTO"]:
            return wb[sheet_name], sheet_name
    
    return None, None

def extract_from_main_sheet(ws_or_wb, filename=None, sheet_name=None):
    """Extrae datos de la pestaña principal (sin RESUMEN)"""
    if hasattr(ws_or_wb, 'active'):
        ws = ws_or_wb.active
    else:
        ws = ws_or_wb
    
    data = []
    
    # Extraer todo el contenido como texto (solo primeras filas importantes)
    all_text = []
    all_cells = {}
    
    for row_idx in range(1, min(30, ws.max_row + 1)):  # Reducir a primeras 30 filas
        for col in range(1, 10):
            val = ws.cell(row_idx, col).value
            if val:
                val_str = str(val).strip()
                all_text.append(val_str.upper())
                all_cells[(row_idx, col)] = (val, val_str)
    
    full_text = " ".join(all_text)
    
    # Detectar tipo: busca "NOTA DE VENTA N°" para identificar notas, no solo "NOTA DE VENTA"
    # Esto evita confundir con la hoja "NOTA DE VENTA" que todos los archivos tienen
    # Si la hoja principal dice claramente "NOTA DE VENTA N°", es una nota
    is_nota_venta = "NOTA DE VENTA N°" in full_text or "NOTA DE VENTA :" in full_text
    
    # Si aún no puede determinar, usa como fallback el nombre de la hoja
    if not is_nota_venta and not ("PRESUPUESTO" in full_text):
        # Si el presupuesto no está explícito y hay NOTA DE VENTA, es nota
        is_nota_venta = "NOTA DE VENTA" in full_text
    
    # Patrones de búsqueda
    cliente = None
    fecha = None
    numero_doc = None
    numero_nv = None
    total = None
    
    # Recorrer primeras filas
    for row_idx in range(1, min(60, ws.max_row + 1)):
        row_vals = []
        for col in range(1, 10):
            val = ws.cell(row_idx, col).value
            if val:
                row_vals.append((col, val))
        
        row_text = " ".join(str(v or "") for _, v in row_vals).upper()
        
        # Buscar cliente (busca después de CLIENTE: o NOMBRE:)
        if not cliente and ("CLIENTE" in row_text or "NOMBRE" in row_text):
            for col_idx in range(len(row_vals)):
                col, val = row_vals[col_idx]
                val_str = str(val).strip()
                # Si encontramos "CLIENTE" o "NOMBRE", el siguiente valor no vacío es el cliente
                if ("CLIENTE" in val_str or "NOMBRE" in val_str):
                    # Buscar en las siguientes columnas el valor del cliente
                    for siguiente_idx in range(col_idx + 1, len(row_vals)):
                        siguiente_val = str(row_vals[siguiente_idx][1]).strip()
                        if siguiente_val and not any(x in siguiente_val.upper() for x in ["CLIENTE", "NOMBRE", "RUT", "PLAZO", "DIAS"]):
                            cliente = siguiente_val
                            break
                    if cliente:
                        break
        
        # Buscar números
        for col, val in row_vals:
            val_str = str(val).strip()
            if val_str.isdigit() and len(val_str) >= 4:
                if not numero_doc:
                    numero_doc = val_str
                elif not numero_nv and is_nota_venta and len(val_str) >= 4 and col > 3:
                    numero_nv = val_str
        
        # Buscar fecha
        if not fecha:
            for col, val in row_vals:
                if isinstance(val, (int, float)) and not isinstance(val, bool):
                    if 40000 < val < 50000:
                        fecha_str = excel_date_to_iso(val)
                        if fecha_str:
                            fecha = fecha_str
                            break
                elif hasattr(val, 'date'):
                    fecha = excel_date_to_iso(val)
                    if fecha:
                        break
    
    # Buscar NETO en CUALQUIER fila (puede variar según cantidad de detalles)
    # Recorre todas las filas hasta encontrar "NETO"
    neto = None
    total = None
    
    for row_idx in range(1, ws.max_row + 1):
        for col in range(1, 8):
            val = ws.cell(row_idx, col).value
            val_str = str(val or "").strip()
            
            # Si encuentra "NETO", busca el número en las siguientes columnas de esta fila
            if "NETO" in val_str.upper() and ":" not in val_str:  # Evita "NETO:" sin valor
                # El valor está en la siguiente columna o en la siguiente fila misma columna
                for siguiente_col in range(col + 1, min(col + 3, 9)):
                    neto_val = ws.cell(row_idx, siguiente_col).value
                    if neto_val and isinstance(neto_val, (int, float)) and neto_val > 1000:
                        neto = float(neto_val)
                        break
                
                # Si no encontró, busca en la siguiente fila misma columna
                if not neto:
                    neto_val = ws.cell(row_idx + 1, col + 1).value
                    if neto_val and isinstance(neto_val, (int, float)) and neto_val > 1000:
                        neto = float(neto_val)
                
                if neto:
                    break
        
        if neto:
            break
    
    # Para BARRANES: no aplica IVA (sin IVA)
    if neto:
        total = neto * 1  # Barranes NO tienen IVA
    
    # Agregar datos
    if cliente and numero_doc and total and total > 0:
        if not fecha:
            fecha = "2026-05-01"
        
        data.append({
            "numero_cotizacion": numero_doc,
            "numero_nota_venta": numero_nv or "",
            "cliente": cliente,
            "fecha": fecha,
            "total": total,
            "tipo": "nota_venta" if is_nota_venta else "cotizacion",
            "canal": "web"
        })
    
    return data if data else None

def process_file(filepath):
    """Procesa un archivo Excel"""
    filepath = Path(filepath)
    
    if not filepath.exists():
        return []
    
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        return []
    
    # Intentar extraer de RESUMEN
    data = extract_from_resumen_sheet(wb, filepath.name)
    
    # Si no hay RESUMEN, extraer de sheet principal
    if not data:
        result = find_data_sheet(wb)
        if result[0] is not None:
            data_sheet, sheet_name = result
            data = extract_from_main_sheet(data_sheet, filepath.name, sheet_name)
    
    wb.close()
    
    return data or []

def main():
    if len(sys.argv) < 2:
        print("Uso: python3 extract_historical_data.py <carpeta> [output.csv]")
        sys.exit(1)
    
    input_path = Path(sys.argv[1])
    output_file = sys.argv[2] if len(sys.argv) > 2 else "datos_extraidos.csv"
    
    all_data = []
    
    # Buscar archivos RECURSIVAMENTE en todas las subcarpetas
    print(f"📁 Buscando archivos Excel en: {input_path}")
    files = sorted(input_path.glob("**/*.xlsx")) + sorted(input_path.glob("**/*.xls"))
    
    if not files:
        print(f"❌ No se encontraron archivos .xlsx o .xls en {input_path}")
        print(f"💡 Verifica que tus archivos estén en subcarpetas dentro de {input_path}")
        sys.exit(1)
    
    print(f"✅ Se encontraron {len(files)} archivo(s)")
    print(f"📁 Procesando...\n")
    
    for idx, filepath in enumerate(files, 1):
        # Mostrar ruta relativa
        rel_path = filepath.relative_to(input_path)
        print(f"  [{idx}/{len(files)}] {rel_path}...", end=" ", flush=True)
        
        data = process_file(filepath)
        if data:
            all_data.extend(data)
            print(f"✅ ({len(data)} registro(s))")
        else:
            print("⚠️ No se extrajeron datos")
    
    # Escribir CSV
    if all_data:
        with open(output_file, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=[
                "numero_cotizacion",
                "numero_nota_venta", 
                "cliente",
                "fecha",
                "total",
                "tipo",
                "canal"
            ])
            writer.writeheader()
            writer.writerows(all_data)
        
        print(f"\n✅ {len(all_data)} registro(s) exportado(s) a: {output_file}")
        
        # Resumen
        notas = [d for d in all_data if d["tipo"] == "nota_venta"]
        cotizaciones = [d for d in all_data if d["tipo"] == "cotizacion"]
        total_ventas = sum(d["total"] for d in all_data)
        
        print(f"\n📊 Resumen:")
        print(f"   Notas de venta: {len(notas)}")
        print(f"   Cotizaciones: {len(cotizaciones)}")
        print(f"   Ingresos totales: ${total_ventas:,.2f}")
        if all_data:
            print(f"   Período: {min(d['fecha'] for d in all_data)} a {max(d['fecha'] for d in all_data)}")
    else:
        print("\n❌ No se extrajeron datos de ningún archivo.")
        sys.exit(1)

if __name__ == "__main__":
    main()
