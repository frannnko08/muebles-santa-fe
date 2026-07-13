#!/usr/bin/env python3
"""
VERSIÓN 3.0 - ULTRA-FLEXIBLE CON VALIDACIÓN
Script ROBUSTO para extraer datos históricos de archivos Excel.

Características:
- Busca cliente/fecha/total con múltiples estrategias
- Valida datos ANTES de guardar (no guarda basura)
- Reporta problemas sin fallar
- Maneja variaciones en estructura sin problemas

Uso:
    python3 extract_historical_data_v3.py /ruta/a/archivos_historicos/ datos_extraidos.csv
"""

import openpyxl
import csv
import sys
from pathlib import Path
from datetime import datetime, timedelta
import re

def excel_date_to_iso(excel_date):
    """Convierte número de fecha Excel a formato ISO (YYYY-MM-DD)"""
    if isinstance(excel_date, str):
        try:
            excel_date = int(float(excel_date))
        except (ValueError, TypeError):
            return None
    
    if isinstance(excel_date, (int, float)):
        try:
            base_date = datetime(1899, 12, 30)
            result_date = base_date + timedelta(days=excel_date)
            return result_date.strftime("%Y-%m-%d")
        except (OverflowError, ValueError):
            return None
    
    if hasattr(excel_date, 'date'):  # datetime object
        return excel_date.strftime("%Y-%m-%d")
    
    return None

def extract_from_resumen_sheet(wb, filename):
    """Extrae datos de la pestaña RESUMEN (si existe)"""
    try:
        ws = wb["RESUMEN"]
    except KeyError:
        return None
    
    data = []
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    
    if len(rows) < 2:
        return None
    
    # Obtener headers (primera fila)
    headers = [str(h or "").strip().upper() for h in rows[0]]
    
    # Detectar tipo de documento
    is_nota_venta = "NOTA DE VENTA" in headers
    
    # Procesar datos (desde fila 2)
    for row in rows[1:]:
        if not any(row):
            continue
        
        try:
            if is_nota_venta:
                # Estructura: COTIZACION | NOTA DE VENTA | NOMBRE | FECHA | TOTAL
                if len(row) >= 5:
                    cotizacion = str(row[0]).strip() if row[0] else ""
                    nota_venta = str(row[1]).strip() if row[1] else ""
                    cliente = str(row[2]).strip() if row[2] else ""
                    fecha_val = row[3]
                    total = row[4]
                    tipo = "nota_venta"
                else:
                    continue
            else:
                # Estructura: N° COTIZACION | CLIENTE | FECHA | TOTAL
                if len(row) >= 4:
                    cotizacion = str(row[0]).strip() if row[0] else ""
                    cliente = str(row[1]).strip() if row[1] else ""
                    fecha_val = row[2]
                    total = row[3]
                    nota_venta = ""
                    tipo = "cotizacion"
                else:
                    continue
            
            # Convertir fecha
            if hasattr(fecha_val, 'date'):  # datetime object
                fecha = fecha_val.strftime("%Y-%m-%d")
            else:
                fecha = excel_date_to_iso(fecha_val)
            
            # Validar y agregar
            total = float(total) if total and isinstance(total, (int, float)) else 0
            if cotizacion and cliente and fecha and total > 0:
                data.append({
                    "numero_cotizacion": cotizacion.replace("N° ", "").strip(),
                    "numero_nota_venta": nota_venta,
                    "cliente": cliente,
                    "fecha": fecha,
                    "total": total,
                    "tipo": tipo,
                    "canal": "web"
                })
        
        except (IndexError, ValueError, TypeError, AttributeError):
            continue
    
    return data if data else None

def find_presupuesto_number(ws):
    """Busca el número de PRESUPUESTO (puede estar en múltiples posiciones)"""
    for row_idx in range(1, min(15, ws.max_row + 1)):
        for col in range(1, 8):
            val = ws.cell(row_idx, col).value
            if val and isinstance(val, str) and "PRESUPUESTO" in val.upper():
                # El número puede estar en la siguiente celda
                next_val = ws.cell(row_idx, col + 1).value
                if next_val and isinstance(next_val, (int, str)):
                    return str(next_val).strip().replace("N° ", "")
    return None

def find_cliente(ws):
    """Busca CLIENTE/NOMBRE en primeras filas (más robusto)"""
    for row_idx in range(1, min(20, ws.max_row + 1)):
        for col in range(1, 8):
            val = ws.cell(row_idx, col).value
            val_str = str(val or "").strip()
            
            # Si encuentra CLIENTE: o NOMBRE:, el siguiente valor es el cliente
            if ("CLIENTE" in val_str.upper() or "NOMBRE" in val_str.upper()) and ":" in val_str:
                # Buscar siguiente valor no vacío
                for siguiente_col in range(col + 1, col + 4):
                    siguiente_val = ws.cell(row_idx, siguiente_col).value
                    siguiente_str = str(siguiente_val or "").strip()
                    if siguiente_str and len(siguiente_str) > 2 and not any(x in siguiente_str.upper() for x in ["CLIENTE", "NOMBRE", "RUT", "PLAZO", "DIAS"]):
                        return siguiente_str
    return None

def find_fecha(ws):
    """Busca FECHA en primeras 20 filas"""
    for row_idx in range(1, min(20, ws.max_row + 1)):
        for col in range(1, 8):
            val = ws.cell(row_idx, col).value
            val_str = str(val or "").strip()
            
            # Si encuentra FECHA:, el siguiente es la fecha
            if ("FECHA" in val_str.upper() or "FECHA" in ws.cell(row_idx, col - 1).value if ws.cell(row_idx, col - 1).value else ""):
                # Buscar número de fecha en esta fila
                for check_col in range(col, min(col + 3, 8)):
                    check_val = ws.cell(row_idx, check_col).value
                    if isinstance(check_val, (int, float)) and 40000 < check_val < 50000:
                        fecha_str = excel_date_to_iso(check_val)
                        if fecha_str:
                            return fecha_str
                    elif hasattr(check_val, 'date'):
                        return check_val.strftime("%Y-%m-%d")
    
    # Fallback: buscar cualquier número de fecha válido
    for row_idx in range(1, min(30, ws.max_row + 1)):
        for col in range(1, 8):
            val = ws.cell(row_idx, col).value
            if isinstance(val, (int, float)) and 40000 < val < 50000:
                fecha_str = excel_date_to_iso(val)
                if fecha_str:
                    return fecha_str
    
    return None

def find_total(ws):
    """Busca TOTAL de forma INTELIGENTE"""
    candidates = []
    
    # Estrategia 1: Buscar línea que dice "TOTAL" y captar número grande después
    for row_idx in range(1, ws.max_row + 1):
        row_text = ""
        for col in range(1, 8):
            val = ws.cell(row_idx, col).value
            row_text += str(val or "") + " "
        
        if "TOTAL" in row_text.upper():
            # Buscar número grande en esta fila
            for col in range(1, 8):
                val = ws.cell(row_idx, col).value
                if isinstance(val, (int, float)) and 30000 < val < 10000000:
                    candidates.append((val, "exact"))
    
    # Estrategia 2: Buscar en últimas 20 filas cualquier número grande
    for row_idx in range(max(1, ws.max_row - 20), ws.max_row + 1):
        for col in range(1, 8):
            val = ws.cell(row_idx, col).value
            if isinstance(val, (int, float)) and 50000 < val < 10000000:
                # Validar que sea razonable (no un costo interno)
                row_text = ""
                for check_col in range(1, 8):
                    row_text += str(ws.cell(row_idx, check_col).value or "") + " "
                
                if not any(x in row_text.upper() for x in ["COSTO", "GASTO", "DEPREC", "INDIREC"]):
                    candidates.append((val, "fallback"))
    
    if candidates:
        # Tomar el más grande y confiable
        candidates.sort(key=lambda x: x[0], reverse=True)
        return float(candidates[0][0])
    
    return None

def extract_from_main_sheet(ws, filename=None):
    """Extrae datos de la hoja principal (CON VALIDACIÓN)"""
    data = []
    
    # Buscar datos clave
    numero_doc = find_presupuesto_number(ws)
    cliente = find_cliente(ws)
    fecha = find_fecha(ws)
    total = find_total(ws)
    
    # Detectar tipo: si ve "NOTA DE VENTA N°" en primeras filas, es nota
    is_nota_venta = False
    for row_idx in range(1, min(30, ws.max_row + 1)):
        row_text = ""
        for col in range(1, 8):
            val = ws.cell(row_idx, col).value
            row_text += str(val or "").upper() + " "
        if "NOTA DE VENTA N°" in row_text or "NOTA DE VENTA :" in row_text:
            is_nota_venta = True
            break
    
    # VALIDAR antes de guardar
    if numero_doc and cliente and fecha and total:
        if total > 0 and total < 100000000:  # Validación de rango sensato
            data.append({
                "numero_cotizacion": numero_doc.replace("N° ", "").strip(),
                "numero_nota_venta": "",
                "cliente": cliente,
                "fecha": fecha,
                "total": total,
                "tipo": "nota_venta" if is_nota_venta else "cotizacion",
                "canal": "web"
            })
    
    return data if data else None

def process_file(filepath):
    """Procesa un archivo Excel (CON MANEJO DE ERRORES)"""
    filepath = Path(filepath)
    
    if not filepath.exists():
        return []
    
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        return []
    
    # Intentar extraer de RESUMEN
    data = extract_from_resumen_sheet(wb, filepath.name)
    
    # Si no hay RESUMEN, extraer de sheet principal
    if not data:
        try:
            # Buscar sheet principal (no especiales)
            for sheet_name in wb.sheetnames:
                if sheet_name not in ["RESUMEN", "REGLAS", "COSTOS", "CUBIERTA", "NOTA DE VENTA", "PRODUCCION", "SEGUIMIENTO"]:
                    ws = wb[sheet_name]
                    data = extract_from_main_sheet(ws, filepath.name)
                    if data:
                        break
        except:
            pass
    
    wb.close()
    
    return data or []

def main():
    if len(sys.argv) < 2:
        print("Uso: python3 extract_historical_data_v3.py <carpeta> [output.csv]")
        sys.exit(1)
    
    input_path = Path(sys.argv[1])
    output_file = sys.argv[2] if len(sys.argv) > 2 else "datos_extraidos.csv"
    
    all_data = []
    success_count = 0
    warn_count = 0
    
    # Buscar archivos RECURSIVAMENTE
    print(f"📁 Buscando archivos Excel en: {input_path}")
    files = sorted(input_path.glob("**/*.xlsx")) + sorted(input_path.glob("**/*.xls"))
    
    if not files:
        print(f"❌ No se encontraron archivos .xlsx o .xls en {input_path}")
        sys.exit(1)
    
    print(f"✅ Se encontraron {len(files)} archivo(s)")
    print(f"📁 Procesando...\n")
    
    for idx, filepath in enumerate(files, 1):
        rel_path = filepath.relative_to(input_path)
        print(f"  [{idx}/{len(files)}] {rel_path}...", end=" ", flush=True)
        
        data = process_file(filepath)
        if data:
            all_data.extend(data)
            success_count += 1
            print(f"✅ ({len(data)} registro(s))")
        else:
            warn_count += 1
            print("⚠️ No extraído")
    
    # Escribir CSV
    if all_data:
        with open(output_file, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=[
                "numero_cotizacion",
                "numero_nota_venta", 
                "cliente",
                "fecha",
                "total",
                "tipo",
                "canal"
            ])
            writer.writeheader()
            writer.writerows(all_data)
        
        print(f"\n✅ {len(all_data)} registro(s) exportado(s) a: {output_file}")
        
        # Resumen
        notas = [d for d in all_data if d["tipo"] == "nota_venta"]
        cotizaciones = [d for d in all_data if d["tipo"] == "cotizacion"]
        total_ventas = sum(d["total"] for d in all_data)
        
        print(f"\n📊 Resumen FINAL:")
        print(f"   Archivos procesados: {success_count}/{len(files)}")
        print(f"   Archivos sin datos: {warn_count}")
        print(f"   ─────────────────────")
        print(f"   Notas de venta: {len(notas)}")
        print(f"   Cotizaciones: {len(cotizaciones)}")
        print(f"   Total ingresos: ${total_ventas:,.2f}")
        if all_data:
            print(f"   Período: {min(d['fecha'] for d in all_data)} a {max(d['fecha'] for d in all_data)}")
        
        # Ratio de sentido
        if len(cotizaciones) > 0:
            ratio = len(notas) / len(cotizaciones)
            print(f"\n   Ratio Notas/Cotizaciones: {ratio:.2f}x")
            if ratio > 2:
                print(f"   ⚠️ Nota: El ratio parece alto. Verifica que sea correcto.")
    else:
        print("\n❌ No se extrajeron datos de ningún archivo.")
        sys.exit(1)

if __name__ == "__main__":
    main()
